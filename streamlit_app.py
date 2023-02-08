#Imports

import transformers
#import datasets

import torch
import pandas as pd
import numpy as np
from transformers import (AutoTokenizer, DistilBertForSequenceClassification)
from transformers.modeling_outputs import SequenceClassifierOutput
import matplotlib.pyplot as plt
import requests
import io
import os
import openpyxl

import streamlit as st

class DistilBertForMultilabelSequenceClassification(DistilBertForSequenceClassification):
    def __init__(self, config):
      super().__init__(config)

    def forward(self,
        input_ids=None,
        attention_mask=None,
        head_mask=None,
        inputs_embeds=None,
        labels=None,
        output_attentions=None,
        output_hidden_states=None,
        return_dict=None):

        return_dict = return_dict if return_dict is not None else self.config.use_return_dict

        outputs = self.distilbert(input_ids,
            attention_mask=attention_mask,
            head_mask=head_mask,
            inputs_embeds=inputs_embeds,
            output_attentions=output_attentions,
            output_hidden_states=output_hidden_states,
            return_dict=return_dict)

        hidden_state = outputs[0]
        pooled_output = hidden_state[:, 0]
        pooled_output = self.dropout(pooled_output)
        logits = self.classifier(pooled_output)

        loss = None
        if labels is not None:
            loss_fct = torch.nn.BCEWithLogitsLoss()
            loss = loss_fct(logits.view(-1, self.num_labels),
                            labels.float().view(-1, self.num_labels))

        if not return_dict:
            output = (logits,) + outputs[2:]
            return ((loss,) + output) if loss is not None else output

        return SequenceClassifierOutput(loss=loss,
            logits=logits,
            hidden_states=outputs.hidden_states,
            attentions=outputs.attentions)

#Setup

st.title("Classification : Virtual Assistant")

@st.cache(allow_output_mutation = True)
def get_model():
    tokenizer = AutoTokenizer.from_pretrained("distilbert-base-uncased")
    model = DistilBertForMultilabelSequenceClassification.from_pretrained("Deopusi/virtual_assistant_classification",use_auth_token = "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",num_labels=14)
    return tokenizer,model


tokenizer,model = get_model()


labels = ['Weather', 'Clock', 'Calendar', 'Map', 'Phone', 'Email', 'Calculator', 'Translator', 'Web search', 'Social media', 'Small talk', 'Message', 'Reminders', 'Music']
path = "BdD1.xlsx"

id2label = {str(i):label for i, label in enumerate(labels)}
label2id = {label:str(i) for i, label in enumerate(labels)}

model.config.id2label = id2label
model.config.label2id = label2id

#Visibility
classified = False

#Input
user_input = st.text_area("Enter sentence to classify :")
button = st.button("Classify")
values = st.checkbox("Show values")
side = ["Weather","Clock","Calendar","Map","Phone","Email","Calculator",\
    "Translator","Web search","Social media","Small talk","Message","Reminders","Music"]


if user_input and button:
    input = torch.tensor([tokenizer(user_input)["input_ids"]])
    logits = model(input)[:2]
    output = torch.nn.Softmax(dim=1)(logits[0])
    output = output[0].tolist()
    result = labels[np.argmax(output)]
    st.write(result)
    classified = True

    st.sidebar.expander('')
    st.sidebar.subheader('Not your wanted answer?')
    choice = st.sidebar.selectbox('Choose your answer',['<select>']+side)
    # confirm = st.sidebar.button('confirm')


    if values:
        y_pos = np.arange(len(labels))
        width = 0.5
        fig, ax = plt.subplots()
        maxl = 0
        hbars = ax.barh(y_pos , output,width, align='center')
        maxl = max(maxl,max(output))
        ax.set_yticks(y_pos, labels=labels)
        ax.invert_yaxis()  # labels read top-to-bottom
        ax.legend()
        # Label with specially formatted floats
        # ax.bar_label(hbars, fmt='%.2f')
        ax.set_xlim(right=min(1,maxl+0.1))  # adjust xlim to fit labels
        st.pyplot(fig)

if classified and choice != '<select>':
    st.write("Your choice has been recorded")
    ind = labels.index(choice)
    vector = ['0']*14
    vector[ind] = '1'
    if not os.path.exists(path):
        df1 = pd.DataFrame(columns=['','text']+labels)
        st.write(df1)
        df1.to_excel(path)
    
    book = openpyxl.load_workbook(path)
    if not 'Sheet2' in book.sheetnames:
        book.create_sheet('Sheet2')
        book.save(path)
    
    writer = pd.ExcelWriter(path, engine = 'openpyxl', if_sheet_exists='overlay', mode = 'a')
    df1 = pd.read_excel(path,sheet_name="Sheet2")
    df1.append(pd.DataFrame(columns =['','text']+labels, data=[[str(df1.shape[0])]+[user_input]+vector]))
    df1.to_excel(writer,sheet_name='Sheet2')
    writer.close()
        

        

